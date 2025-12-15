# data_exporter.py - Moduł eksportu danych

import os
import json
import csv
import time
from datetime import datetime
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

class DataExporter:
    """Eksportuje dane sygnałów do różnych formatów"""
    
    def __init__(self, export_dir='exports'):
        self.export_dir = export_dir
        Path(export_dir).mkdir(exist_ok=True)
        self._last_export_time = None
    
    def _generate_filename(self, extension):
        """Generuj unikalną nazwę pliku"""
        # Ensure unique timestamp by adding delay if necessary
        current_time = datetime.now()
        if self._last_export_time and current_time == self._last_export_time:
            time.sleep(0.01)  # Small delay to ensure different timestamp
            current_time = datetime.now()
        
        self._last_export_time = current_time
        timestamp = current_time.strftime('%Y%m%d_%H%M%S')
        return f"modbus_data_{timestamp}.{extension}"
    
    def export_to_csv(self, signals, filename=None):
        """
        Eksportuj dane do CSV
        
        Args:
            signals: Lista słowników z danymi sygnałów
            filename: Nazwa pliku (opcjonalnie)
        
        Returns:
            str: Ścieżka do pliku
        """
        try:
            if not filename:
                filename = self._generate_filename('csv')
            
            filepath = os.path.join(self.export_dir, filename)
            
            if not signals:
                logger.warning("Brak danych do eksportu")
                # Utwórz pusty plik CSV z nagłówkami
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    fieldnames = ['ID', 'Adres', 'Nazwa Sygnału', 'Wartość', 'Jednostka', 'Status', 'Ostatni Odczyt']
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                return filepath
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                fieldnames = ['ID', 'Adres', 'Nazwa Sygnału', 'Wartość', 'Jednostka', 'Status', 'Ostatni Odczyt']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                
                writer.writeheader()
                for signal in signals:
                    writer.writerow({
                        'ID': signal.get('id', ''),
                        'Adres': signal.get('address', ''),
                        'Nazwa Sygnału': signal.get('name', ''),
                        'Wartość': signal.get('value', ''),
                        'Jednostka': signal.get('unit', ''),
                        'Status': signal.get('status', ''),
                        'Ostatni Odczyt': signal.get('lastUpdate', '')
                    })
            
            logger.info(f"✓ Wyeksportowano do CSV: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Błąd eksportu CSV: {str(e)}")
            raise
    
    def export_to_excel(self, signals, filename=None):
        """
        Eksportuj dane do Excel
        
        Args:
            signals: Lista słowników z danymi sygnałów
            filename: Nazwa pliku (opcjonalnie)
        
        Returns:
            str: Ścieżka do pliku
        """
        try:
            if not filename:
                filename = self._generate_filename('xlsx')
            
            filepath = os.path.join(self.export_dir, filename)
            
            # Import openpyxl tylko jeśli jest dostępny
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                logger.error("openpyxl nie zainstalowany. Zainstaluj: pip install openpyxl")
                raise
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sygnały"
            
            # Nagłówki
            headers = ['ID', 'Adres', 'Nazwa Sygnału', 'Wartość', 'Jednostka', 'Status', 'Ostatni Odczyt']
            ws.append(headers)
            
            # Style dla nagłówków
            header_fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if signals:
                # Dane
                for signal in signals:
                    ws.append([
                        signal.get('id', ''),
                        signal.get('address', ''),
                        signal.get('name', ''),
                        signal.get('value', ''),
                        signal.get('unit', ''),
                        signal.get('status', ''),
                        signal.get('lastUpdate', '')
                    ])
            else:
                logger.warning("Brak danych do eksportu")
            
            # Szerokość kolumn
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 20
            
            wb.save(filepath)
            logger.info(f"✓ Wyeksportowano do Excel: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Błąd eksportu Excel: {str(e)}")
            raise
    
    def export_to_json(self, signals, filename=None):
        """
        Eksportuj dane do JSON
        
        Args:
            signals: Lista słowników z danymi sygnałów
            filename: Nazwa pliku (opcjonalnie)
        
        Returns:
            str: Ścieżka do pliku
        """
        try:
            if not filename:
                filename = self._generate_filename('json')
            
            filepath = os.path.join(self.export_dir, filename)
            
            data = {
                'exportDate': datetime.now().isoformat(),
                'signalCount': len(signals),
                'signals': signals
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"✓ Wyeksportowano do JSON: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Błąd eksportu JSON: {str(e)}")
            raise
    
    def export_all(self, signals):
        """
        Eksportuj do wszystkich formatów naraz
        
        Args:
            signals: Lista słowników z danymi sygnałów
        
        Returns:
            dict: Słownik z nazwami wyeksportowanych plików
        """
        try:
            files = {}
            
            csv_file = self.export_to_csv(signals)
            files['csv'] = os.path.basename(csv_file)
            
            json_file = self.export_to_json(signals)
            files['json'] = os.path.basename(json_file)
            
            try:
                excel_file = self.export_to_excel(signals)
                files['excel'] = os.path.basename(excel_file)
            except:
                logger.warning("Nie udało się wyeksportować do Excel")
            
            return files
            
        except Exception as e:
            logger.error(f"Błąd eksportu: {str(e)}")
            raise
